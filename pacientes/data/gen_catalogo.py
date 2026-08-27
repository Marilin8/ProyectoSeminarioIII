"""Lee el tarifario Excel y genera un catálogo normalizado de estudios
con modalidad, duración y matriz de precios por convenio × horario.

Salida: catalogo_estudios.json  (lista de dicts)
"""
import json
import re
import sys

import openpyxl

XLSX = r"C:\Users\elmer\Downloads\PRECIOS GENERAL .xlsx"

# --- Modalidades -------------------------------------------------------------
RX = "rx"
RX_CONTRASTE = "rx_contraste"
TAC = "tac"
USG = "usg"
MAMO_DENSIT = "mamo_densit"

# Estudios de RX que en realidad son fluoroscopía / con medio de contraste /
# procedimiento (van en otra sala, requieren preparación y más tiempo).
KW_RX_CONTRASTE = (
    "SERIE", "ENEMA", "ESOFAGOGRAMA", "HISTEROSALPINGO", "HSG", "URETROGRAMA",
    "CISTOGRAMA", "CISTOMICC", "PIELOGRAMA", "FISTULOGRAMA", "COLANGIOGRAMA",
    "ILEOSTOGRAMA", "COLOSTOGRAMA", "UROGRAMA",
)


def normalizar_nombre(raw, modalidad_prefijo):
    s = str(raw).strip().upper()
    # quita el prefijo de modalidad redundante: "RX.", "TAC.", "USG.", "RX "
    s = re.sub(r"^(RX|TAC|USG)\.?\s+", "", s)
    # errores de tipeo puntuales observados en el archivo
    s = s.replace("RAXOS", "RAYOS").replace("ADOMEN", "ABDOMEN")
    s = s.replace("URETRETROGRAMA", "URETROGRAMA")
    # acrónimos con puntos -> forma compacta (antes de borrar puntos sueltos)
    s = re.sub(r"S\s*\.?\s*G\s*\.?\s*D\s*\.?", "SGD", s)
    s = re.sub(r"S\s*\.?\s*G\s*\.?\s*I\s*\.?", "SGI", s)
    # expande abreviaturas para poder deduplicar entre hojas
    s = re.sub(r"\bC\.\s+CERVICAL\b", "COLUMNA CERVICAL", s)
    s = re.sub(r"\bCOL\.?\s+", "COLUMNA ", s)
    s = re.sub(r"\bART\.?\s+", "ARTICULACION ", s)
    s = re.sub(r"\bTEM\.?\s*MAND\.?", "TEMPOROMANDIBULAR", s)
    s = re.sub(r"\bAP\s*/\s*LATERAL\b", "AP Y LATERAL", s)
    # quita cualquier punto suelto que haya quedado en medio del nombre
    s = s.replace(".", " ")
    s = re.sub(r"\s*/\s*", " / ", s)        # normaliza separador "/"
    s = re.sub(r"\(\s+", "(", s)
    s = re.sub(r"\s+\)", ")", s)
    s = re.sub(r"\s{2,}", " ", s)           # espacios repetidos
    s = s.strip(" /-")
    # la pielotomografía siempre es contrastada: unifica los dos nombres
    if s.startswith("PIELOTOMOGRAFIA"):
        s = "PIELOTOMOGRAFIA"
    return s


def duracion_para(modalidad, nombre):
    n = nombre.upper()
    if modalidad == RX_CONTRASTE:
        return 45
    if modalidad == RX:
        if any(k in n for k in ("COLUMNA", "PELVIS", "DINAMIC", "FLEXION", "OBLICU",
                                "SACRO", "PELVIMETRIA", "ESCANOGRAMA", "CRANEO", "SENOS")):
            return 20
        return 10
    if modalidad == TAC:
        if any(k in n for k in ("CONTRASTADA", "ABDOMEN COMPLETO", "TRIFASICA",
                                "ANGIO", "UROTOMOGRAFIA", "PIELOTOMOGRAFIA")):
            return 75
        return 45
    if modalidad == USG:
        return 45 if "DOPPLER" in n else 30
    if modalidad == MAMO_DENSIT:
        return 30
    return 30


def num(v):
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    return None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    catalogo = {}   # nombre_normalizado -> dict

    def registrar(nombre, modalidad, **precios):
        item = catalogo.setdefault(nombre, {
            "nombre": nombre,
            "modalidad": modalidad,
            "duracion_minutos": duracion_para(modalidad, nombre),
            "precios": {},
        })
        # si un mismo estudio aparece como RX simple y RX contraste, gana contraste
        if modalidad == RX_CONTRASTE:
            item["modalidad"] = RX_CONTRASTE
            item["duracion_minutos"] = duracion_para(RX_CONTRASTE, nombre)
        for k, v in precios.items():
            if v is not None:
                item["precios"][k] = v

    # ---- PRIVADO: RAYOS X (col G=hábil, H=inhábil) ----
    ws = wb["RAYOS X PRIV"]
    seccion = RX
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if not isinstance(a, str):
            continue
        a = a.strip()
        if not a or a.startswith("LISTADO") or a == "ESTUDIOS":
            continue
        if a == "RAYOS X":
            seccion = RX
            continue
        if a.startswith("MAMOGRAFIA") and "ULTRASONIDO" in a:
            seccion = MAMO_DENSIT
            continue
        if a.upper().startswith(("USO DEL EQUIPO", "PLACA EXTRA")):
            continue  # cargos de equipo, no estudios
        if seccion == MAMO_DENSIT:
            if a.upper().startswith("COMBO"):
                continue  # los combos se manejan aparte (punto 5, no ahora)
            precio_h = num(row[6])
            registrar(normalizar_nombre(a, "MAMO"), MAMO_DENSIT,
                      privado_habil=precio_h, privado_inhabil=precio_h)
            continue
        # sección Rayos X
        habil, inhabil = num(row[6]), num(row[7])
        mod = RX_CONTRASTE if any(k in a.upper() for k in KW_RX_CONTRASTE) else RX
        registrar(normalizar_nombre(a, "RX"), mod,
                  privado_habil=habil,
                  privado_inhabil=inhabil if inhabil is not None else habil)

    # ---- PRIVADO: TAC (col G=hábil, H=inhábil) ----
    ws = wb["TAC PRIV"]
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if not isinstance(a, str):
            continue
        a = a.strip()
        if (not a or a.startswith("LISTADO") or a.startswith("ESTUDIOS")
                or a.isupper() and num(row[6]) is None):
            continue
        habil, inhabil = num(row[6]), num(row[7])
        if habil is None:
            continue
        registrar(normalizar_nombre(a, "TAC"), TAC,
                  privado_habil=habil,
                  privado_inhabil=inhabil if inhabil is not None else habil)

    # ---- PRIVADO: USG (G=hábil, H=inhábil, I=precio social) ----
    ws = wb["USG "]
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if not isinstance(a, str):
            continue
        a = a.strip()
        if not a or a.startswith("LISTADO") or a == "ESTUDIOS" or a == "ULTRASONIDO":
            continue
        if a.upper() == "PREPARACION" or "(" in a:
            continue  # notas de preparación
        habil, inhabil, social = num(row[6]), num(row[7]), num(row[8])
        if habil is None:
            continue
        registrar(normalizar_nombre(a, "USG"), USG,
                  privado_habil=habil,
                  privado_inhabil=inhabil if inhabil is not None else habil,
                  coex_habil=social)

    # ---- COEX: RAYOS X desde hoja EPSS (un solo precio, hábil) ----
    ws = wb["EPSS"]
    for row in ws.iter_rows(values_only=True):
        a, precio = row[0], num(row[1])
        if not isinstance(a, str):
            continue
        a = a.strip()
        if not a or precio is None or a.upper() in ("RAYOS X", "PRECIO", "NOMBRE DEL ESTUDIO"):
            continue
        if "ELECTROCARDIOGRAMA" in a.upper():
            continue  # no es imagenología
        mod = RX_CONTRASTE if any(k in a.upper() for k in KW_RX_CONTRASTE) else RX
        if a.upper().startswith(("MAMOGRAFIA", "DENSITOMETRIA")):
            mod = MAMO_DENSIT
        nombre = normalizar_nombre(a, "RX" if mod != MAMO_DENSIT else "MAMO")
        registrar(nombre, mod, coex_habil=precio)

    # ---- COEX: TAC desde hoja SOCIAL (solo tarifa hábil; COEX no tiene inhábil) ----
    ws = wb[" SOCIAL "]
    for row in ws.iter_rows(values_only=True):
        a = row[0]
        if not isinstance(a, str):
            continue
        a = a.strip()
        habil = num(row[6])
        if not a or habil is None:
            continue
        registrar(normalizar_nombre(a, "TAC"), TAC, coex_habil=habil)

    # ---- Completa huecos ----
    # COEX no tiene tarifa inhábil (solo hábil). EMERGENCIA IGSS y el COEX que
    # no tenga tarifa propia -> se asume la privada como punto de partida
    # (ajustable después desde el admin).
    for item in catalogo.values():
        p = item["precios"]
        if "privado_habil" in p:
            p.setdefault("coex_habil", p["privado_habil"])
            p.setdefault("emergencia_igss_habil", p["privado_habil"])
        if "privado_inhabil" in p:
            p.setdefault("emergencia_igss_inhabil", p["privado_inhabil"])
        # si falta la variante inhábil en privado/emergencia, = la hábil
        for conv in ("privado", "emergencia_igss"):
            h, ih = f"{conv}_habil", f"{conv}_inhabil"
            if h in p and ih not in p:
                p[ih] = p[h]

    salida = sorted(catalogo.values(), key=lambda x: (x["modalidad"], x["nombre"]))
    with open("catalogo_estudios.json", "w", encoding="utf-8") as fh:
        json.dump(salida, fh, ensure_ascii=False, indent=2)

    # resumen
    from collections import Counter
    c = Counter(i["modalidad"] for i in salida)
    print("Total estudios normalizados:", len(salida))
    for k, v in sorted(c.items()):
        print(f"  {k:14} {v}")
    sin_privado = [i["nombre"] for i in salida if "privado_habil" not in i["precios"]]
    sin_coex = [i["nombre"] for i in salida if "coex_habil" not in i["precios"]]
    print("Sin precio privado:", len(sin_privado))
    print("Sin precio coex:", len(sin_coex))
    print()
    for i in salida:
        print(f'{i["modalidad"]:13} | {i["duracion_minutos"]:>3}m | {i["nombre"]}')
        print(f'                     precios: {i["precios"]}')


if __name__ == "__main__":
    sys.exit(main())
